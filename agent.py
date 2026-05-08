"""
Agente PDF → Excel usando Agno + Mistral
Escritório de Contabilidade — Extração inteligente de dados de PDFs
"""

import json
import os
import re
from pathlib import Path

from dotenv import load_dotenv

# Carrega variáveis do .env (se existir)
load_dotenv()


def _carregar_api_key() -> str | None:
    """
    Carrega a API key do Mistral com a seguinte prioridade:
    1. Variável de ambiente MISTRAL_API_KEY (inclui .env)
    2. Arquivo api.txt na raiz do projeto
    """
    # 1. Variável de ambiente (já carregada pelo dotenv se existir .env)
    chave = os.environ.get("MISTRAL_API_KEY")
    if chave:
        return chave.strip()

    # 2. Fallback: arquivo api.txt
    arquivo_api = Path(__file__).parent / "api.txt"
    if arquivo_api.exists():
        linhas = arquivo_api.read_text(encoding="utf-8").strip().splitlines()
        # A chave é a última linha não vazia
        for linha in reversed(linhas):
            linha = linha.strip()
            if linha and linha.lower() != "mistral":
                os.environ["MISTRAL_API_KEY"] = linha
                return linha

    return None


# Carrega a API key ao importar o módulo
MISTRAL_API_KEY = _carregar_api_key()

import pdfplumber
from agno.agent import Agent
from agno.models.mistral import MistralChat
from agno.tools import tool
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ──────────────────────────────────────────────
# FERRAMENTAS DO AGENTE
# ──────────────────────────────────────────────

@tool
def extrair_texto_pdf(caminho_pdf: str) -> str:
    """
    Extrai o texto completo de um arquivo PDF, página por página.
    Retorna o conteúdo como string formatada para análise.
    """
    if not Path(caminho_pdf).exists():
        return f"ERRO: Arquivo não encontrado: {caminho_pdf}"

    texto_completo = []
    with pdfplumber.open(caminho_pdf) as pdf:
        for i, pagina in enumerate(pdf.pages, 1):
            texto = pagina.extract_text()
            if texto:
                texto_completo.append(f"=== PÁGINA {i} ===\n{texto}")

            # Tenta extrair tabelas também
            tabelas = pagina.extract_tables()
            for j, tabela in enumerate(tabelas, 1):
                linhas_formatadas = []
                for linha in tabela:
                    linha_limpa = [str(c) if c else "" for c in linha]
                    linhas_formatadas.append(" | ".join(linha_limpa))
                texto_completo.append(f"--- Tabela {j} (Pág {i}) ---\n" + "\n".join(linhas_formatadas))

    texto_final = "\n\n".join(texto_completo) if texto_completo else "AVISO: Nenhum texto extraível encontrado no PDF."
    
    # Limite seguro de caracteres para evitar o erro de contexto máximo (262k tokens do Mistral)
    # 800.000 caracteres equivalem a aproximadamente 200.000 tokens.
    limite_caracteres = 800_000
    if len(texto_final) > limite_caracteres:
        texto_final = texto_final[:limite_caracteres]
        texto_final += "\n\n[AVISO DO SISTEMA: O PDF é extremamente grande e teve seu conteúdo cortado para caber no limite de memória da IA. Alguns dados do final do arquivo podem não ser extraídos.]"
        
    return texto_final


@tool
def identificar_tipo_documento(texto_pdf: str) -> str:
    """
    Analisa o texto extraído e identifica o tipo de documento contábil/fiscal.
    Retorna o tipo identificado e os campos esperados.
    """
    texto_lower = texto_pdf.lower()

    tipos = {
        "nota_fiscal": ["nota fiscal", "nf-e", "chave de acesso", "cfop", "icms", "cnpj emitente"],
        "boleto": ["boleto", "linha digitável", "beneficiário", "nosso número", "vencimento", "valor do documento"],
        "extrato_bancario": ["extrato", "saldo", "débito", "crédito", "agência", "conta corrente"],
        "balancete": ["balancete", "débitos", "créditos", "saldo devedor", "saldo credor"],
        "folha_pagamento": ["folha de pagamento", "salário", "inss", "irrf", "fgts", "vencimentos", "descontos"],
        "recibo": ["recibo", "recebi", "valor recebido", "pagador"],
        "contrato": ["contrato", "contratante", "contratado", "cláusula"],
        "dfe": ["dfe", "escrituração", "sped"],
    }

    tipo_detectado = "documento_generico"
    maior_score = 0

    for tipo, palavras_chave in tipos.items():
        score = sum(1 for p in palavras_chave if p in texto_lower)
        if score > maior_score:
            maior_score = score
            tipo_detectado = tipo

    campos_esperados = {
        "nota_fiscal": ["numero_nf", "data_emissao", "emitente_cnpj", "emitente_nome", "destinatario_cnpj",
                        "destinatario_nome", "valor_total", "valor_icms", "valor_ipi", "cfop", "descricao_produto"],
        "boleto": ["beneficiario", "pagador", "valor", "vencimento", "nosso_numero", "linha_digitavel", "banco"],
        "extrato_bancario": ["data", "descricao", "valor", "tipo_lancamento", "saldo"],
        "balancete": ["conta", "descricao", "saldo_anterior", "debitos", "creditos", "saldo_atual"],
        "folha_pagamento": ["nome_funcionario", "cargo", "salario_base", "inss", "irrf", "fgts",
                            "outros_descontos", "valor_liquido"],
        "recibo": ["pagador", "beneficiario", "valor", "descricao", "data"],
        "contrato": ["contratante", "contratado", "objeto", "valor", "prazo", "data_assinatura"],
        "documento_generico": ["campo1", "campo2", "campo3", "valor", "data", "descricao"],
    }

    return json.dumps({
        "tipo": tipo_detectado,
        "campos_esperados": campos_esperados.get(tipo_detectado, []),
        "confianca": f"{min(maior_score * 20, 100)}%"
    }, ensure_ascii=False)


@tool
def salvar_excel(dados_json: str, caminho_saida: str, titulo_planilha: str = "Dados Extraídos") -> str:
    """
    Salva os dados extraídos em um arquivo Excel formatado profissionalmente.
    dados_json deve ser uma lista de dicionários JSON.
    """
    try:
        dados = json.loads(dados_json)
        if not dados:
            return "ERRO: Nenhum dado para salvar."

        # Garante que é lista
        if isinstance(dados, dict):
            dados = [dados]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = titulo_planilha[:31]  # Excel limita a 31 chars

        # Estilos
        cor_cabecalho = "1F4E79"
        cor_linha_par = "D6E4F0"
        fonte_cabecalho = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        fonte_dados = Font(name="Arial", size=10)
        alinhamento_centro = Alignment(horizontal="center", vertical="center")
        alinhamento_esquerda = Alignment(horizontal="left", vertical="center", wrap_text=True)
        borda = Border(
            left=Side(style="thin", color="BFBFBF"),
            right=Side(style="thin", color="BFBFBF"),
            top=Side(style="thin", color="BFBFBF"),
            bottom=Side(style="thin", color="BFBFBF"),
        )

        # Cabeçalho de título
        colunas = list(dados[0].keys())
        ws.merge_cells(f"A1:{get_column_letter(len(colunas))}1")
        celula_titulo = ws["A1"]
        celula_titulo.value = titulo_planilha
        celula_titulo.font = Font(name="Arial", bold=True, color="FFFFFF", size=13)
        celula_titulo.fill = PatternFill("solid", fgColor="0D2137")
        celula_titulo.alignment = alinhamento_centro
        ws.row_dimensions[1].height = 28

        # Cabeçalhos das colunas (linha 2)
        for col_idx, coluna in enumerate(colunas, 1):
            celula = ws.cell(row=2, column=col_idx, value=coluna.replace("_", " ").title())
            celula.font = fonte_cabecalho
            celula.fill = PatternFill("solid", fgColor=cor_cabecalho)
            celula.alignment = alinhamento_centro
            celula.border = borda
        ws.row_dimensions[2].height = 22

        # Dados
        for linha_idx, registro in enumerate(dados, 3):
            for col_idx, coluna in enumerate(colunas, 1):
                valor = registro.get(coluna, "")
                celula = ws.cell(row=linha_idx, column=col_idx, value=valor)
                celula.font = fonte_dados
                celula.border = borda
                celula.alignment = alinhamento_esquerda

                # Linhas alternadas
                if linha_idx % 2 == 0:
                    celula.fill = PatternFill("solid", fgColor=cor_linha_par)

            ws.row_dimensions[linha_idx].height = 18

        # Ajusta largura das colunas automaticamente
        for col_idx, coluna in enumerate(colunas, 1):
            max_largura = len(coluna) + 4
            for linha_idx in range(3, len(dados) + 3):
                valor = ws.cell(row=linha_idx, column=col_idx).value
                if valor:
                    max_largura = max(max_largura, min(len(str(valor)) + 2, 50))
            ws.column_dimensions[get_column_letter(col_idx)].width = max_largura

        # Congela cabeçalhos
        ws.freeze_panes = "A3"

        # Adiciona linha de total se houver campos numéricos
        ultima_linha = len(dados) + 3
        for col_idx, coluna in enumerate(colunas, 1):
            valores = []
            for r in dados:
                v = r.get(coluna, "")
                try:
                    # Tenta converter para número (remove R$, pontos, etc.)
                    v_limpo = re.sub(r"[R$\s\.]", "", str(v)).replace(",", ".")
                    valores.append(float(v_limpo))
                except (ValueError, AttributeError):
                    pass

            if valores and len(valores) == len(dados):
                col_letra = get_column_letter(col_idx)
                celula_total = ws.cell(row=ultima_linha, column=col_idx,
                                       value=f"=SUM({col_letra}3:{col_letra}{ultima_linha - 1})")
                celula_total.font = Font(name="Arial", bold=True, size=10)
                celula_total.fill = PatternFill("solid", fgColor="1F4E79")
                celula_total.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
                celula_total.border = borda
                celula_total.alignment = alinhamento_centro
            elif col_idx == 1:
                celula_total = ws.cell(row=ultima_linha, column=col_idx, value="TOTAL")
                celula_total.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
                celula_total.fill = PatternFill("solid", fgColor="1F4E79")
                celula_total.border = borda
                celula_total.alignment = alinhamento_esquerda

        Path(caminho_saida).parent.mkdir(parents=True, exist_ok=True)
        wb.save(caminho_saida)
        return f"✅ Excel salvo com sucesso: {caminho_saida} ({len(dados)} registros)"

    except json.JSONDecodeError as e:
        return f"ERRO ao parsear JSON: {e}"
    except Exception as e:
        return f"ERRO ao salvar Excel: {e}"


# ──────────────────────────────────────────────
# CRIAÇÃO DO AGENTE
# ──────────────────────────────────────────────

def criar_agente(campos_desejados: list[str] | None = None) -> Agent:
    """
    Cria e retorna o agente Agno com Mistral configurado.
    campos_desejados: lista de campos específicos que o usuário quer extrair.
    """
    campos_instrucao = ""
    if campos_desejados:
        campos_instrucao = f"""
O usuário quer extrair ESPECIFICAMENTE estes campos:
{json.dumps(campos_desejados, ensure_ascii=False, indent=2)}

Foque em extrair exatamente esses campos. Se um campo não for encontrado, use null.
"""

    instrucoes = f"""Você é um agente especialista em análise de documentos contábeis e fiscais brasileiros.

Sua missão é:
1. Extrair texto do PDF usando a ferramenta `extrair_texto_pdf`
2. Identificar o tipo de documento usando `identificar_tipo_documento`
3. Analisar o conteúdo extraído com inteligência e precisão
4. Estruturar os dados encontrados em formato JSON (lista de dicionários)
5. Salvar o resultado em Excel usando `salvar_excel`

{campos_instrucao}

REGRAS IMPORTANTES:
- Sempre normalize valores monetários para formato numérico (ex: "R$ 1.234,56" → "1234.56")
- Datas devem estar no formato DD/MM/AAAA
- CNPJs e CPFs devem preservar a formatação original
- Se o documento tiver múltiplos registros (ex: itens de NF, lançamentos de extrato), crie uma linha por registro
- Se não encontrar um campo, use string vazia "" — nunca invente dados
- Seja preciso e fiel ao conteúdo do documento

Após salvar o Excel, informe ao usuário:
- Tipo de documento identificado
- Quantidade de registros extraídos  
- Campos extraídos
- Caminho do arquivo gerado
"""

    agente = Agent(
        model=MistralChat(
            id="mistral-large-latest",
            api_key=MISTRAL_API_KEY,
            client_params={"timeout_ms": 1_800_000},  # 30 minutos
        ),
        tools=[extrair_texto_pdf, identificar_tipo_documento, salvar_excel],
        instructions=instrucoes,
        markdown=True,
    )

    return agente


# ──────────────────────────────────────────────
# FUNÇÃO PRINCIPAL
# ──────────────────────────────────────────────

def processar_pdf(
    caminho_pdf: str,
    caminho_saida: str | None = None,
    campos_desejados: list[str] | None = None,
) -> str:
    """
    Processa um PDF e gera um Excel com os dados extraídos.

    Args:
        caminho_pdf: Caminho para o arquivo PDF
        caminho_saida: Caminho para salvar o Excel (opcional, gera automaticamente)
        campos_desejados: Lista de campos específicos a extrair (opcional)

    Returns:
        Resposta do agente com o resultado da operação
    """
    if not caminho_saida:
        nome_base = Path(caminho_pdf).stem
        caminho_saida = str(Path(caminho_pdf).parent / f"{nome_base}_extraido.xlsx")

    agente = criar_agente(campos_desejados)

    prompt = f"""
Processe o seguinte arquivo PDF e extraia os dados para Excel:

- **PDF**: `{caminho_pdf}`
- **Salvar em**: `{caminho_saida}`

Siga o processo:
1. Extraia o texto do PDF
2. Identifique o tipo de documento
3. Extraia todos os dados relevantes estruturados
4. Salve no Excel com o título adequado ao tipo de documento
"""

    if campos_desejados:
        prompt += f"\n\nExtraia especificamente estes campos: {', '.join(campos_desejados)}"

    resposta = agente.run(prompt)
    return resposta.content


# ──────────────────────────────────────────────
# EXECUÇÃO DIRETA
# ──────────────────────────────────────────────

if __name__ == "__main__":
    import sys

    print("=" * 60)
    print("  AGENTE PDF → EXCEL | Escritório de Contabilidade")
    print("  Powered by Agno + Mistral")
    print("=" * 60)

    # Verifica API Key
    if not MISTRAL_API_KEY:
        print("\n⚠️  API Key não encontrada!")
        print("   Opções:")
        print("   1. Crie um arquivo .env com: MISTRAL_API_KEY=sua-chave")
        print("   2. Coloque a chave no arquivo api.txt")
        print("   3. Defina a variável de ambiente: $env:MISTRAL_API_KEY='sua-chave'")
        sys.exit(1)

    # Uso via linha de comando
    if len(sys.argv) < 2:
        print("\nUso: python agent.py <caminho_pdf> [campo1,campo2,...] [saida.xlsx]")
        print("\nExemplos:")
        print("  python agent.py nota_fiscal.pdf")
        print("  python agent.py extrato.pdf data,descricao,valor,saldo")
        print("  python agent.py boleto.pdf beneficiario,valor,vencimento resultado.xlsx")
        sys.exit(1)

    pdf = sys.argv[1]
    campos = sys.argv[2].split(",") if len(sys.argv) > 2 and not sys.argv[2].endswith(".xlsx") else None
    saida = sys.argv[-1] if sys.argv[-1].endswith(".xlsx") else None

    print(f"\n📄 Processando: {pdf}")
    if campos:
        print(f"🎯 Campos desejados: {', '.join(campos)}")

    resultado = processar_pdf(pdf, saida, campos)
    print("\n" + resultado)